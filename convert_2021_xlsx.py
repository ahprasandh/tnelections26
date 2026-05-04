#!/usr/bin/env python3
"""Convert all 2021 TN election XLSX files to JSON."""

import json
import os
import re
import openpyxl
from openpyxl.reader.excel import ExcelReader

XLSX_DIR = "2021"
OUT_DIR = "2021"


def load_workbook_safe(path):
    """Load workbook, handling stylesheet errors."""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except ValueError:
        wb = openpyxl.Workbook()
        reader = ExcelReader(path, read_only=True, data_only=True)
        reader.read_manifest()
        reader.read_strings()
        reader.read_workbook()
        reader.read_worksheets()
        return reader.wb


def get_rows(ws):
    """Get all rows as lists from a worksheet."""
    return [list(r) for r in ws.iter_rows(values_only=True)]


def save_json(data, filename):
    path = os.path.join(OUT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  -> {path} ({len(json.dumps(data)):,} bytes)")


def convert_abbreviations():
    """1- Other Abbreviations and Description"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "1- Other Abbreviations and Description.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = []
    for row in rows[1:]:  # skip title row
        abbr, _, desc = row[0], row[1], row[2]
        if abbr and desc:
            result.append({"abbreviation": str(abbr).strip(), "description": str(desc).strip()})
    wb.close()
    save_json(result, "abbreviations.json")


def convert_successful_candidates():
    """2- List of Successful Candidates"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "2- List of Successful Candidates.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    # Header row is row[1]: STATE/UT, CONSTITUENCY, WINNER, SEX, PARTY, SYMBOL
    result = []
    for row in rows[2:]:  # skip title + header
        if row[1]:
            result.append({
                "constituency": str(row[1]).strip(),
                "winner": str(row[2]).strip(),
                "sex": str(row[3]).strip(),
                "party": str(row[4]).strip(),
                "symbol": str(row[5]).strip() if row[5] else ""
            })
    wb.close()
    save_json(result, "successful_candidates.json")


def convert_parties_participated():
    """3- List of Political Parties Participated"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "3- List of Political Parties Participated.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = {"national_parties": [], "state_parties": [], "registered_unrecognized_parties": []}
    current_type = None
    for row in rows[2:]:  # skip title + header
        val = row[0]
        if val and isinstance(val, str) and not val.replace(".", "").isdigit():
            key = val.strip().lower().replace(" ", "_")
            if "national" in key:
                current_type = "national_parties"
            elif "state" in key:
                current_type = "state_parties"
            elif "registered" in key or "unrecognized" in key:
                current_type = "registered_unrecognized_parties"
            continue
        if row[1] and current_type:
            result[current_type].append({
                "abbreviation": str(row[1]).strip(),
                "party_name": str(row[2]).strip()
            })
    wb.close()
    save_json(result, "parties_participated.json")


def convert_highlights():
    """4- Highlights"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "4- Highlights.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = {}
    i = 0
    while i < len(rows):
        row = rows[i]
        if row[0] and isinstance(row[0], (int, float)):
            label = str(row[1]).strip() if row[1] else ""
            # Check if next rows have sub-items
            if label.lower().startswith("no. of constituencies"):
                # Next row has Type of Constituency header, then data
                if i + 2 < len(rows):
                    data_row = rows[i + 2]
                    result["constituencies"] = {
                        "general": data_row[2],
                        "sc": data_row[3],
                        "st": data_row[4],
                        "total": data_row[5]
                    }
                    i += 3
                    continue
            elif label.lower().startswith("no") and "contest" in label.lower():
                sub = {}
                i += 1
                while i < len(rows) and rows[i][0] is None and rows[i][1]:
                    sub_label = str(rows[i][1]).strip().rstrip(".")
                    sub[sub_label] = rows[i][2] if len(rows[i]) > 2 else None
                    i += 1
                result["contestants"] = sub
                continue
            else:
                # Generic key-value
                result[label] = row[2] if len(row) > 2 and row[2] is not None else True
        i += 1
    wb.close()
    save_json(result, "highlights.json")


def convert_party_performance():
    """5- Performance of Political Parties"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "5- Performance of Political Parties.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = {"national_parties": [], "state_parties": [], "registered_unrecognized_parties": [], "independents": None}
    current_type = None
    for row in rows[3:]:  # skip title + 2 header rows
        val = row[0]
        if val and isinstance(val, str):
            key = val.strip().lower()
            if "national" in key:
                current_type = "national_parties"
                continue
            elif "state" in key:
                current_type = "state_parties"
                continue
            elif "registered" in key or "unrecognized" in key:
                current_type = "registered_unrecognized_parties"
                continue
            elif "independent" in key:
                current_type = "independents"
                continue
            elif "total" in key:
                current_type = None
                continue
        if row[1] and current_type and current_type != "independents":
            entry = {
                "abbreviation": str(row[1]).strip(),
                "seats_contested": row[2],
                "seats_won": row[3],
                "forfeited_deposits": row[4],
                "votes": row[5],
                "vote_share_pct": str(row[6]).strip() if row[6] else "0%",
            }
            result[current_type].append(entry)
        elif current_type == "independents" and row[2] is not None:
            result["independents"] = {
                "seats_contested": row[2],
                "seats_won": row[3],
                "forfeited_deposits": row[4],
                "votes": row[5],
                "vote_share_pct": str(row[6]).strip() if row[6] else "0%",
            }
    wb.close()
    save_json(result, "party_performance.json")


def convert_electors_data():
    """6- Electors Data Summary"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "6- Electors Data Summary.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = {}
    for row in rows[3:]:  # skip title + header rows
        if row[0] is not None and row[1]:
            label = str(row[1]).strip() if row[1] else str(row[0]).strip()
            result[label] = {
                "general": row[2],
                "sc": row[3],
                "st": row[4],
                "total": row[5]
            }
        elif row[0] is None and row[1]:
            label = str(row[1]).strip()
            result[label] = {
                "general": row[2],
                "sc": row[3],
                "st": row[4],
                "total": row[5]
            }
    wb.close()
    save_json(result, "electors_data_summary.json")


def convert_women_candidates():
    """7- Individual Performance of Women Candidates"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "7- Individual Performance of Women Candidates.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = []
    current_constituency = None
    for row in rows:
        # Constituency header rows
        if row[0] and isinstance(row[0], str) and "constituency" in row[0].lower():
            continue
        if row[0] and isinstance(row[0], str) and ("sl" in row[0].lower() or "name of" in row[0].lower()):
            continue
        # Check for constituency name in merged cells
        if row[0] and isinstance(row[0], str) and not row[0].replace(".", "").strip().isdigit():
            # Could be a constituency name or a header
            if row[1] is None and row[2] is None:
                current_constituency = str(row[0]).strip()
                continue
        # Data rows have sl.no as first column
        if row[0] and isinstance(row[0], (int, float)) and row[1]:
            entry = {
                "constituency": current_constituency,
                "sl_no": int(row[0]),
                "name": str(row[1]).strip(),
                "party": str(row[2]).strip() if row[2] else "",
                "party_type": str(row[3]).strip() if row[3] else "",
                "votes_secured": row[4],
                "pct_over_electors": row[5],
                "pct_over_votes_polled": row[6],
                "status": str(row[7]).strip() if row[7] else "",
                "total_valid_votes": row[8]
            }
            result.append(entry)
    wb.close()
    save_json(result, "women_candidates.json")


def convert_constituency_data():
    """8- Constituency Data Summary - each sheet is a constituency"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "8- Constituency Data Summary.xlsx"))
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = get_rows(ws)
        if len(rows) < 3:
            continue
        # Row 1: header, Row 2: State/UT info with constituency name
        const_info = str(rows[1][3]).strip() if rows[1][3] else sheet_name
        # Parse constituency number and name
        match = re.match(r"(\d+)-(.+?)-(GEN|SC|ST)", const_info)
        if match:
            const_no = int(match.group(1))
            const_name = match.group(2).strip()
            const_type = match.group(3)
        else:
            const_no = None
            const_name = const_info
            const_type = ""

        constituency = {
            "const_no": const_no,
            "constituency": const_name,
            "type": const_type,
            "candidates": {},
            "electors": {},
            "voting": {}
        }

        section = None
        for row in rows[2:]:
            label = str(row[0]).strip() if row[0] else ""
            sub_label = str(row[1]).strip() if row[1] else ""

            if "Candidates" in label:
                section = "candidates"
                continue
            elif "Electors" in label:
                section = "electors"
                continue
            elif "Voting" in label or "Votes" in label:
                section = "voting"
                continue

            if sub_label and section:
                key = sub_label.strip().rstrip(".")
                # Clean number prefix
                key = re.sub(r"^\d+\.\s*", "", key)
                vals = {"men": row[2], "women": row[3], "third_gender": row[4], "total": row[5]}
                if section == "electors":
                    vals = {"general": row[2], "sc": row[3], "st": row[4] if len(row) > 4 else None, "total": row[5] if len(row) > 5 else row[4]}
                constituency[section][key] = vals

        result.append(constituency)
    wb.close()
    save_json(result, "constituency_data_summary.json")


def convert_candidates_data_summary():
    """9- Candidates Data Summary"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "9- Candidates Data Summary.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = {}
    for row in rows[4:]:  # skip title + headers
        label = str(row[0]).strip() if row[0] else ""
        if label:
            result[label] = {
                "general": row[1],
                "sc": row[2],
                "st": row[3],
                "total": row[4]
            }
    wb.close()
    save_json(result, "candidates_data_summary.json")


def convert_detailed_results():
    """10- Detailed Results - the most important file"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "10- Detailed Results.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    # Header at row[3]: STATE/UT NAME, AC NO., AC NAME, CANDIDATE NAME, SEX, AGE, CATEGORY, PARTY, SYMBOL, GENERAL, POSTAL, TOTAL, % VOTES POLLED, TOTAL ELECTORS
    constituencies = {}
    for row in rows[4:]:  # skip title + header rows
        if not row[1] or not row[3]:
            continue
        ac_no = int(row[1]) if row[1] else None
        ac_name = str(row[2]).strip() if row[2] else ""
        # Clean candidate name (remove leading number)
        cand_name = str(row[3]).strip()
        cand_name = re.sub(r"^\d+\s+", "", cand_name)

        candidate = {
            "name": cand_name,
            "sex": str(row[4]).strip() if row[4] else "",
            "age": int(row[5]) if row[5] else None,
            "category": str(row[6]).strip() if row[6] else "",
            "party": str(row[7]).strip() if row[7] else "",
            "symbol": str(row[8]).strip() if row[8] else "",
            "votes_general": int(row[9]) if row[9] is not None else 0,
            "votes_postal": int(row[10]) if row[10] is not None else 0,
            "votes_total": int(row[11]) if row[11] is not None else 0,
            "vote_share_pct": float(row[12]) if row[12] is not None else 0.0,
        }

        if ac_no not in constituencies:
            constituencies[ac_no] = {
                "const_no": ac_no,
                "constituency": ac_name,
                "total_electors": int(row[13]) if row[13] is not None else 0,
                "candidates": []
            }
        constituencies[ac_no]["candidates"].append(candidate)

    # Sort candidates within each constituency by total votes (winner first)
    for c in constituencies.values():
        c["candidates"].sort(key=lambda x: x["votes_total"], reverse=True)

    result = {"constituencies": sorted(constituencies.values(), key=lambda x: x["const_no"])}
    wb.close()
    save_json(result, "detailed_results.json")


def convert_annexure():
    """Annexure-1 (Electors Data Summary)"""
    wb = load_workbook_safe(os.path.join(XLSX_DIR, "Annexure-1.xlsx"))
    ws = wb.active
    rows = get_rows(ws)
    result = {}
    for row in rows[3:]:
        label = str(row[0]).strip() if row[0] else ""
        if label and row[1] is not None:
            result[label] = {
                "general": row[1],
                "sc": row[2],
                "st": row[3],
                "total": row[4]
            }
        elif not label and row[0] is None and row[1] is not None:
            # Sub-items
            sub_label = str(row[1]).strip() if isinstance(row[1], str) else str(row[1])
            if sub_label:
                result[sub_label] = {
                    "general": row[2] if len(row) > 2 else None,
                    "sc": row[3] if len(row) > 3 else None,
                    "st": row[4] if len(row) > 4 else None,
                    "total": row[5] if len(row) > 5 else None
                }
    wb.close()
    save_json(result, "annexure_electors_data.json")


if __name__ == "__main__":
    print("Converting 2021 TN Election XLSX files to JSON...")
    print()

    print("1. Abbreviations")
    convert_abbreviations()

    print("2. Successful Candidates")
    convert_successful_candidates()

    print("3. Parties Participated")
    convert_parties_participated()

    print("4. Highlights")
    convert_highlights()

    print("5. Party Performance")
    convert_party_performance()

    print("6. Electors Data Summary")
    convert_electors_data()

    print("7. Women Candidates")
    convert_women_candidates()

    print("8. Constituency Data Summary")
    convert_constituency_data()

    print("9. Candidates Data Summary")
    convert_candidates_data_summary()

    print("10. Detailed Results")
    convert_detailed_results()

    print("11. Annexure-1")
    convert_annexure()

    print()
    print("Done! All files converted.")
